"""
This script combines a folder of excel files into one file.

It places a header row, and then pulls the specified cells from each of those files
into a new row for each file. Puts file name in first cell.
Author: Jordan Welch
"""
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.write_only import WriteOnlyCell
import glob
import Tkinter as tk
import tkFileDialog
from openpyxl.writer.write_only import WriteOnlyCell

'''
CONSTANTS - EDIT CONSTANTS HERE
'''


#Top row of the excel file being generated
HEADER_ROW = ["BOM", "OH Passings", "UG Passings"]

#Cells that should be pulled from each file in the input folder.
FROM_CELLS = [
[6,5],[6,6]
]

'''
END CONSTANTS
'''

#Hide the TK base window so it doesn't confuse
root = tk.Tk()
root.withdraw()


input_dir = tkFileDialog.askdirectory(parent=root,initialdir="~",title='Select Input Directory')
if len(input_dir ) > 0:
	print "Input Directory" + input_dir
else:
	sys.exit()
	
input_files = glob.glob(input_dir + "\*.xlsx")

row_to = 1

#Create Output File
wb_output = Workbook(write_only=True)
ws_output = wb_output.create_sheet()

#Write header
temp_header = []
for header in HEADER_ROW:
	temp_header.append(WriteOnlyCell(ws_output, value=header))

ws_output.append(temp_header)

for file in input_files:
	#GET NAME OF FILE
	path_split = file.split("/")
	file_name = path_split[len(path_split) - 1]
	file_name = file_name.split("\\")[1]
	
	wb_input = load_workbook(file, read_only=True, data_only=True)
	
	ws_input = wb_input.active
	
	row_to_add = []
	
	#Tile of row
	row_to_add.append(WriteOnlyCell(ws_output, value=file_name.split("_")[0]))
	
	for cell_from in FROM_CELLS:
		row_to_add.append(WriteOnlyCell(ws_output, value=(ws_input.cell(row=cell_from[0],column=cell_from[1]).value)))
	
	ws_output.append(row_to_add)

print("Saving File")
wb_output.save("./OUTPUT.XLSX")		